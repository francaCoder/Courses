"""
- Print on screen just the value 'Philip' of "pessoas.xlsx"
- Change this value to "Jimmy"
- Print on screen the values of lines 2 and 10
- Print on screen all values inside the 'Last name' column
"""

import openpyxl

spreadsheet = openpyxl.load_workbook("8 - pessoas.xlsx")
print(spreadsheet.sheetnames)
sheet1 = spreadsheet.get_sheet_by_name('Sheet1')

#1
print(sheet1["B4"].value)

#2
sheet1["B4"].value = "Jimmy"

#3
for line in sheet1.iter_rows(min_row=2, max_row=11):
    print(line[0].value, line[1].value, line[2].value, line[3].value, line[4].value, line[5].value, line[6].value, line[7].value, )

#4
for line in sheet1.iter_rows(min_col=3, max_col=3):
    for cell in line:
        print(cell.value)

# Sheets = Page
# Workbook = File with pages

import openpyxl

spreadsheet_test = openpyxl.Workbook()
spreadsheet_test.create_sheet("Fruits")
spreadsheet_test.create_sheet("Vegetables")
spreadsheet_test.create_sheet("Seeds")
print(spreadsheet_test.sheetnames)
header_fruits = ["Title", "Localization", "Price"]
sheet_fruits = spreadsheet_test.get_sheet_by_name("Fruits")
sheet_fruits.append(header_fruits)
spreadsheet_test.save("Supermarket Data.xlsx")

fruits = [["Grape", "Market", 5], ["Watermelon", "Market", 15], ["Cake", "Market", 40]]

for fruit in fruits:
    sheet_fruits.append(fruit)

spreadsheet_test.save("Supermarket Data.xlsx")

import openpyxl

spreadsheet = openpyxl.load_workbook("8 - pessoas.xlsx") # Open spreadsheet
print(spreadsheet.sheetnames) # What are the pages - Sheets
sheet1 = spreadsheet.get_sheet_by_name('Sheet1') # To specify
print(sheet1["C3"].value) # Name - Hashimoto / Access a field

# If I can change
sheet1["C3"].value = "França"
print(sheet1["C3"].value) # Name - França

# Or

sheet1.cell(row=3, column=3, value="Matheus")
print(sheet1["C3"].value) # Name - França

print()

# Sweep a spreadsheet
for line in sheet1.iter_rows(min_row=2, max_row=10, min_col=2, max_col=4):
    print(line[0].value, line[1].value, line[2].value)

# Just one column
for line in sheet1.iter_rows(min_col=2, max_col=2):
    for cell in line:
        print(cell.value)
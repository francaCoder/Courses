from time import sleep
from selenium import webdriver
import openpyxl


class SearchPrices:
    def __init__(self):
        self.driver = webdriver.Chrome(executable_path="chromedriver.exe")

    def start(self):
        self.current_page_number = 1
        self.driver.get(f"https://mg.olx.com.br/belo-horizonte-e-regiao?q=computadores")
        self.create_worksheet()
        self.find_values_in_page()

    def create_worksheet(self):
        self.worksheet = openpyxl.Workbook()
        self.worksheet.create_sheet("values")
        self.worksheet_values = self.worksheet["values"]
        # self.worksheet_name = "Computer Prices.xlsx"
        # self.worksheet_values = self.worksheet.active
        # self.worksheet_values.title = "Computer Prices"
        self.worksheet_values.cell(row=1, column=1, value="Title")
        self.worksheet_values.cell(row=1, column=2, value="Locations")
        self.worksheet_values.cell(row=1, column=3, value="Price")

    def find_values_in_page(self):
        try:
            while True:
                self.titles = self.driver.find_elements_by_xpath("//h2[@class='sc-1mbetcw-0 fKteoJ sc-ifAKCX jyXVpA']")

                self.locations = self.driver.find_elements_by_xpath("//span[@class=sc-7l84qu-1 ciykCV sc-ifAKCX dpURtf']")

                self.prices = self.driver.find_elements_by_xpath("//span[@class='sc-ifAKCX eoKYee']")

                self.save_values_in_worksheet()
                self.next_page()
        except Exception as error:
            print("no more searches to extract information")
            print(error)

    def next_page(self):
        self.current_page_number += 1
        # sleep(1)
        self.driver.get(f"https://mg.olx.com.br/belo-horizonte-e-regiao?o={self.current_page_number}&q=computadores")

    def save_values_in_worksheet(self):
        for i in range(0, len(self.titles)):
            new_line = [self.titles[i].text, self.prices[i].text]
            self.worksheet_values.append(new_line)
        self.worksheet.save("Computer prices BH.xlsx")


extract = SearchPrices()
extract.start()


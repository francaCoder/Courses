import openpyxl
worksheet = openpyxl.Workbook()
worksheet.create_sheet("Fruits")
worksheet.create_sheet("Vegetables")
worksheet.create_sheet("Seeds")
print(worksheet.sheetnames)

header_fruits = ["Title", "Localization", "Price"]
fruits_sheet = worksheet.get_sheet_by_name("Fruits")
fruits_sheet.append(header_fruits)

# Change name
# window_name.title = "new name"

worksheet.save("Super Market Data.xlsx")
